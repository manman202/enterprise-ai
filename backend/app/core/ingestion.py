"""
Document ingestion pipeline — extract text → chunk → embed → store in ChromaDB.
Supports: .txt, .pdf, .docx, .xlsx
"""

import hashlib
import logging
import os
import re
from datetime import datetime

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.embeddings import embed_batch
from app.db.chroma import add_document_chunks
from app.models.document import Document

logger = logging.getLogger(__name__)

# Chunking configuration
CHUNK_SIZE = 500  # Target characters per chunk
CHUNK_OVERLAP = 50  # Characters shared between adjacent chunks


# ── Text Extraction ────────────────────────────────────────────────────────────


def extract_text_from_txt(filepath: str) -> str:
    """Read plain text file."""
    with open(filepath, encoding="utf-8", errors="replace") as f:
        return f.read()


def extract_text_from_pdf(filepath: str) -> str:
    """Extract text from PDF using pypdf."""
    try:
        from pypdf import PdfReader  # type: ignore[import]

        reader = PdfReader(filepath)
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n\n".join(pages)
    except ImportError:
        logger.warning("pypdf not installed — cannot extract PDF text")
        return ""
    except Exception as exc:
        logger.warning("PDF extraction failed for '%s': %s", filepath, exc)
        return ""


def extract_text_from_docx(filepath: str) -> str:
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document as DocxDocument  # type: ignore[import]

        doc = DocxDocument(filepath)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except ImportError:
        logger.warning("python-docx not installed — cannot extract DOCX text")
        return ""
    except Exception as exc:
        logger.warning("DOCX extraction failed for '%s': %s", filepath, exc)
        return ""


def extract_text_from_xlsx(filepath: str) -> str:
    """Extract text from XLSX — each cell on its own line, sheets separated."""
    try:
        import openpyxl  # type: ignore[import]

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                row_cells = [str(cell) for cell in row if cell is not None]
                if row_cells:
                    lines.append(" | ".join(row_cells))
        return "\n".join(lines)
    except ImportError:
        logger.warning("openpyxl not installed — cannot extract XLSX text")
        return ""
    except Exception as exc:
        logger.warning("XLSX extraction failed for '%s': %s", filepath, exc)
        return ""


def extract_text(filepath: str) -> str:
    """
    Dispatch text extraction based on file extension.
    Returns empty string if the format is unsupported.
    """
    ext = os.path.splitext(filepath)[1].lower()
    dispatch = {
        ".txt": extract_text_from_txt,
        ".md": extract_text_from_txt,
        ".pdf": extract_text_from_pdf,
        ".docx": extract_text_from_docx,
        ".xlsx": extract_text_from_xlsx,
        ".xls": extract_text_from_xlsx,
    }
    extractor = dispatch.get(ext)
    if extractor is None:
        logger.warning("Unsupported file type '%s' — skipping text extraction", ext)
        return ""
    return extractor(filepath)


# ── Chunking ───────────────────────────────────────────────────────────────────


def chunk_text(text: str, size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    """
    Split text into overlapping chunks on sentence/paragraph boundaries.
    Preserves semantic context across chunk boundaries via overlap.

    Args:
        text: Full document text
        size: Target chunk size in characters
        overlap: Characters of overlap between consecutive chunks

    Returns:
        List of text chunks
    """
    # Normalize whitespace
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    if not text:
        return []

    chunks: list[str] = []
    start = 0

    while start < len(text):
        end = start + size

        if end >= len(text):
            # Last chunk — take everything remaining
            chunks.append(text[start:].strip())
            break

        # Try to break at a sentence boundary (. ! ?) or newline
        boundary = -1
        for sep in ["\n\n", "\n", ". ", "! ", "? "]:
            idx = text.rfind(sep, start, end)
            if idx > start:
                boundary = idx + len(sep)
                break

        if boundary == -1:
            # No good boundary — break at word boundary
            space_idx = text.rfind(" ", start, end)
            boundary = space_idx + 1 if space_idx > start else end

        chunk = text[start:boundary].strip()
        if chunk:
            chunks.append(chunk)

        # Move forward with overlap
        start = max(boundary - overlap, start + 1)

    return chunks


# ── MD5 hash ───────────────────────────────────────────────────────────────────


def file_md5(content: bytes) -> str:
    """Compute MD5 hash of file content bytes."""
    return hashlib.md5(content).hexdigest()


# ── Full Ingestion Pipeline ────────────────────────────────────────────────────


async def ingest_document(
    doc: Document,
    content: bytes,
    db: AsyncSession,
) -> None:
    """
    Run the full ingestion pipeline for a document:
    1. Compute file hash (for dedup detection)
    2. Extract text by file type
    3. Chunk the text
    4. Embed all chunks in a batch
    5. Store chunks + embeddings in ChromaDB
    6. Update Document record in DB

    Errors are caught and stored in doc.error_message — they do not propagate.
    """
    # Mark as in-progress
    doc.status = "pending"
    doc.file_hash = file_md5(content)
    doc.size = len(content)
    await db.commit()

    try:
        # Step 1: extract text
        # Write to a temp file if content is in memory (avoid holding large bytes)
        import tempfile

        ext = os.path.splitext(doc.filename)[1].lower()
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        try:
            text = extract_text(tmp_path)
        finally:
            os.unlink(tmp_path)  # Always clean up temp file

        if not text.strip():
            raise ValueError("No text could be extracted from the document")

        # Step 2: chunk
        chunks = chunk_text(text)
        if not chunks:
            raise ValueError("Chunking produced no output")

        # Step 3: embed (batch for efficiency)
        embeddings = embed_batch(chunks)

        # Step 4: store in ChromaDB
        await add_document_chunks(
            doc_id=doc.id,
            chunks=chunks,
            embeddings=embeddings,
            filename=doc.filename,
            department=doc.department or "",
        )

        # Step 5: mark as ingested
        doc.status = "ingested"
        doc.chunks_count = len(chunks)
        doc.ingested_at = datetime.utcnow()
        doc.error_message = None
        logger.info(
            "Document '%s' ingested: %d chunks from %d chars",
            doc.filename,
            len(chunks),
            len(text),
        )

    except Exception as exc:  # noqa: BLE001
        doc.status = "failed"
        doc.error_message = str(exc)
        logger.error("Ingestion failed for '%s': %s", doc.filename, exc)

    await db.commit()
